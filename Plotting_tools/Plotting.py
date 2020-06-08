from __future__ import division
import numpy as np
np.seterr(divide='ignore', invalid='ignore')
import time
import os
class Plotting(object):
    def __init__(self):
        self.plotspec = [[] for _ in range(6)]
        self.timearrays = []
        self.dataarrays = []
        self.filenames = []
        self.offsets = []
        self.timeoffset = [0]
        self.scales = []
        self.xlimit = []
        self.ylimits = [[] for _ in range(6)]
        self.channel_names = []
        self.xlabels = [[] for _ in range(6)]
        self.ylabels = [[] for _ in range(6)]
        self.titles = [[] for _ in range(6)]
        self.legends = [[] for _ in range(6)]
        self.rstimes = [[] for _ in range(6)]
        self.damping_cals = [[] for _ in range(6)]
        self.plot_iqs = [[] for _ in range(6)]
        self.plot_Vdroops = [[] for _ in range(6)]
        self.plot_Qdroops = [[] for _ in range(6)]
        self.plot_Vdroops_Hybrids = [[] for _ in range(6)]
        self.P_Recoverys = [[] for _ in range(6)]
        self.plot_PFs = [[] for _ in range(6)]
        self.checkspikes = [[] for _ in range(6)]

    def plot_spike(self,tDataArray,outDataArray, thereshold, type):
        import matplotlib.pyplot as plt
        dataPTS = len(tDataArray)
        timeSubSet = tDataArray[:dataPTS]
        QSubSet = outDataArray[:dataPTS]
        ymax = np.amax(QSubSet)
        maximum = 0
        if ymax>=thereshold:
            for i, value in enumerate(QSubSet):
                if value > maximum:
                    maximum = value
                    time_max = tDataArray[i]
            t1 = time_max - 0.5  # this function apply for if there is no other spike in 500ms before, and the duration of the spike start to max less than 500ms
            if type == 'zoom_in':
                self.xlimit = [int(time_max) - 0.5, int(time_max) + 1.5]
            else:
                pass
            QSubSetIndex = 0
            time = 0
            while time <= t1:
                QSubSetIndex += 1
                time = timeSubSet[QSubSetIndex]
                Qini = QSubSet[QSubSetIndex - 1]
            while Qini < thereshold:
                QSubSetIndex += 1
                Qini = QSubSet[QSubSetIndex - 1]
            T_start = timeSubSet[QSubSetIndex-1]
            while Qini >= thereshold:
                QSubSetIndex += 1
                Qini = QSubSet[QSubSetIndex - 1]
            T_stop = timeSubSet[QSubSetIndex-1]

            ymin, ymax = plt.ylim()
            xmin, xmax = plt.xlim()

            plt.plot([T_start, T_start], [ymin, ymax], linestyle='--', color='k', linewidth=1.2)
            plt.plot([T_stop, T_stop], [ymin, ymax], linestyle='--', color='k', linewidth=1.2)
            plt.plot([xmin, xmax], [thereshold, thereshold], linestyle=':', color='r', linewidth=1.2)

            plt.autoscale(enable=True, axis='both', tight=False)
            duration = abs(T_stop - T_start)*1000
            ymax = np.amax(QSubSet)
            ymin = np.amin(QSubSet)
            xtext = xmin + 0.1 * (xmax - xmin)
            ytext1 = ymin + 0.6 *(ymax - ymin)
            ytext2 = ymin + 0.5 *(ymax-ymin)
            ytext3 = ymin + 0.4 *(ymax - ymin)
            # plt.text(xtext, ytext, r'Spike Duration:  %.1f (ms), Vmax: %.3f (pu)' % (duration,ymax), color='blue',fontsize=14)
            # plt.text(xtext, ytext2, r'Start from %1.4f (s) to %1.4f (s)' %(T_start,T_stop),color='blue',fontsize=14)
            if type == 'zoom_in':
                plt.text(int(time_max) - 0.4, thereshold *1.01, r'Threshold:  %.3f (pu)' % (thereshold),color='r', fontsize=14)
                plt.text(int(time_max) - 0.4, ytext1, r'Spike (over %.3f p.u) Duration:  %.1f (ms)' % (thereshold, duration), color='blue',fontsize=14)
                plt.text(int(time_max) - 0.4, ytext2, r'Start from %1.4f (s) to %1.4f (s)' %(T_start,T_stop),color='blue',fontsize=14)
                plt.text(int(time_max) - 0.4, ytext3, r'Maximum: %.3f (pu)' % (ymax), color='blue',fontsize=14)
            else:
                plt.text(xtext, thereshold *1.01, r'Threshold:  %.3f (pu)' % (thereshold),color='r', fontsize=14)
                plt.text(xtext, ytext1, r'Spike (over %.3f p.u) Duration:  %.1f (ms)' % (thereshold, duration), color='blue',fontsize=14)
                plt.text(xtext, ytext2, r'Start from %1.4f (s) to %1.4f (s)' %(T_start,T_stop),color='blue',fontsize=14)
                plt.text(xtext, ytext3, r'Maximum: %.3f (pu)' % (ymax), color='blue',fontsize=14)

    def PF_plot(self,tDataArray,P_POC, Q_POC, PF_set):
        import matplotlib.pyplot as plt
        import math
        t = tDataArray
        tanpi = PF_set
        S_POC = []
        PF_ref = []
        PF = []
        for i in range(len(t)):
            Si = ((P_POC[i]**2)+(Q_POC[i]**2))**0.5
            if Si == 0:
                Si = 0.001
            S_POC.append(Si)
        for i in range(len(t)):
            PFi =P_POC[i]/S_POC[i] * np.sign(Q_POC[i])# math.copysign(P_pre[i]/S_pre[i],Q_pre[i])
            PF.append(PFi)
            PF_ref_i = math.cos(math.atan(tanpi[i]))* np.sign(Q_POC[i])
            PF_ref.append(PF_ref_i)
        plt.plot(t,PF,label="POC_PF",linewidth = 3.5) #change ax1. to plt.
        plt.plot(t,PF_ref,label="POC_PF_ref",linestyle="dashed", dashes=(4, 3),linewidth = 3.5)
        plt.autoscale(enable=True, axis='both', tight=False)

    # def rise_settle_TimeCalc(self,tDataArray,outDataArray,rs_xlimit,rs_ylimit,tFault,Tfin,filename,workbook):
    def rise_settle_TimeCalc(self,tDataArray,outDataArray,rs_xlimit,rs_ylimit,tFault,Tfin):
        import matplotlib.pyplot as plt
        dataPTS = len(tDataArray)
        # timeSubSet = tDataArray[:dataPTS/1]
        # QSubSet = outDataArray[:dataPTS/1]

        timeSubSet = tDataArray[:dataPTS]
        QSubSet = outDataArray[:dataPTS]

        QSubSetIndex = 0
        time = 0
        while time <= tFault:
            QSubSetIndex += 1
            time = timeSubSet[QSubSetIndex]
        Qini = QSubSet[QSubSetIndex-1]
        while time <= Tfin:
            QSubSetIndex += 1
            time = timeSubSet[QSubSetIndex]
        Qfin = QSubSet[QSubSetIndex-1]
        QSubSetIndexFin = QSubSetIndex
        Qchange = abs(Qfin-Qini)

        QSubSetIndex = 0
        x = 0

        #while x <= 0.005*Qchange:    # --> x <= 0.001*Qchange/Qchange
        #    QSubSetIndex += 1
        #    x = abs(QSubSet[QSubSetIndex]-Qini)
        #T_ini = timeSubSet[QSubSetIndex]
        T_ini = tFault
        # if (isinstance(tFault, float)) or (isinstance(tFault, list)):
        #     T_ini = tFault# Make it easy, the time that we change reference signal.
        # elif (isinstance(tFault, str)):
        #     while x <= 0.005*Qchange:    # --> x <= 0.001*Qchange/Qchange
        #         QSubSetIndex += 1
        #         x = abs(QSubSet[QSubSetIndex]-Qini)
        #     T_ini = timeSubSet[QSubSetIndex]

        QSubSetIndex = 0
        time = 0
        while time <= T_ini:
            time = timeSubSet[QSubSetIndex]
            QSubSetIndex += 1
        x = abs(QSubSet[QSubSetIndex] - Qini)

        while x <= 0.1*Qchange:
            QSubSetIndex += 1
            x = abs(QSubSet[QSubSetIndex]-Qini)
        T_10pc = timeSubSet[QSubSetIndex]

        while x <= 0.9*Qchange:
            QSubSetIndex += 1
            x = abs(QSubSet[QSubSetIndex]-Qini)
        T_90pc = timeSubSet[QSubSetIndex]

        QSubSetIndex = QSubSetIndexFin
        x = 0
        while x < 0.1*Qchange:
            QSubSetIndex -= 1
            x = abs(QSubSet[QSubSetIndex]-Qfin)
        T_setl = timeSubSet[QSubSetIndex]

        ymin, ymax = plt.ylim()
        xmin, xmax = plt.xlim()

        plt.plot([T_10pc,T_10pc],[ymin,ymax],linestyle='-', color='k', linewidth=1.2)
        plt.plot([T_90pc,T_90pc],[ymin,ymax],linestyle='-', color='k', linewidth=1.2)

        plt.plot([T_ini,T_ini],[ymin,ymax],linestyle='--', color='r', linewidth=1.2)
        plt.plot([T_setl,T_setl],[ymin,ymax],linestyle='--', color='r', linewidth=1.2)

        plt.autoscale(enable=True, axis='both', tight=False)
    
        riseTime = abs(T_90pc-T_10pc)
        setlTime = abs(T_setl-T_ini)

        ymax = np.amax(QSubSet)
        ymin = np.amin(QSubSet)

        if rs_xlimit != []:
            xmin, xmax = rs_xlimit
        if rs_ylimit != []:
            ymin, ymax = rs_ylimit

        xtext = xmin + 0.5 *(xmax-xmin)
        ytext = ymin + 0.5 *(ymax-ymin)
        ytext2 = ymin + 0.35 *(ymax-ymin)
        plt.text(xtext, ytext, r'Rise time:  %.3f s, Settling time: %.3f s'%(riseTime,setlTime),color='blue',fontsize=14)
        plt.text(xtext, ytext2, r'Change of %1.3f from %1.3f to %1.3f' %(Qchange,Qini,Qfin),color='blue',fontsize=14)

        # if workbook != "" and filename != "":
        #
        #     from openpyxl import load_workbook
        #     wb = load_workbook(workbook)
        #     ws = wb.active
        #     row = ws.max_row
        #     if ws.cell(row=row, column=4).value != None:
        #         ws.cell(row=row+1, column=1).value = os.path.basename(os.path.dirname(filename))
        #         ws.cell(row=row+1, column=2).value = riseTime
        #         ws.cell(row=row+1, column=3).value = setlTime
        #     else:
        #         ws.cell(row=row, column=4).value = riseTime
        #         ws.cell(row=row, column=5).value = setlTime
        #     wb.save(workbook)


    def damping_calculation(self, tDataArray, outDataArray, rs_xlimit,rs_ylimit, Parameter, value):
        import numpy as np
        import math
        import matplotlib.pyplot as plt
        from scipy.signal import find_peaks
        from statistics import mean
        Time = tDataArray
        x = outDataArray
        final_average_value = round(mean(x.tolist()[-100:]), 10)
        if Parameter == 'distance':
            peaks, _ = find_peaks(x, distance=value)
        elif Parameter == 'width':
            peaks, _ = find_peaks(x, width=value)
        elif Parameter == 'threshold':
            peaks, _ = find_peaks(x, threshold=value)
        else:
            peaks, _ = find_peaks(x, prominence=value)

        plt.plot(Time[peaks], x[peaks], 'vr', markersize=12)
        plt.autoscale(enable=True, axis='both', tight=False)

        t_peaks = Time[peaks].tolist()
        Y_peaks = x[peaks].tolist()
        periods = []
        for i in range(len(t_peaks) - 1):
            period = t_peaks[i + 1] - t_peaks[i]
            periods.append(period)
        Average = round(mean(periods), 10)
        frequency = round(1 / Average, 10)
        log_dec = 1 / (len(Y_peaks) - 1) * math.log(abs(Y_peaks[0] / Y_peaks[-1]))
        damping_ratio = 1 / math.sqrt((1 + (2 * math.pi / log_dec) ** 2))

        Y_peak_max = Y_peaks[0]
        t_peak_max = t_peaks[0]
        for i in range(len(Y_peaks)):
            if Y_peaks[i] > Y_peak_max:
                Y_peak_max = Y_peaks[i]
                t_peak_max = t_peaks[i]
        ys = []
        ts = []
        for i in range(len(t_peaks)):
            if t_peaks[i] >= t_peak_max:
                ys.append(Y_peaks[i])
                ts.append(t_peaks[i])

        ys = np.array(ys)
        ts = np.array(ts)

        y1 = (0.5 * (Y_peak_max - final_average_value) + final_average_value)
        y2 = (0.25 * (Y_peak_max - final_average_value) + final_average_value)
        y3 = (0.125 * (Y_peak_max - final_average_value) + final_average_value)
        y4 = (0.0625 * (Y_peak_max - final_average_value) + final_average_value)

        print(ys)
        print(ts)

        t1 = np.interp(y1, ys, ts)
        t2 = np.interp(y2, ys, ts)
        t3 = np.interp(y3, ys, ts)
        t4 = np.interp(y4, ys, ts)
        halving1 = t1 - t_peak_max
        halving2 = t2 - t1
        halving3 = t3 - t2
        halving4 = t4 - t3
        halving_time = halving1

        print(Y_peak_max, final_average_value)
        print(ys)
        print(ts)
        print(t1,t2,t3,t4)
        print(y1,y2,y3,y4)

        dataPTS = len(tDataArray)
        QSubSet = outDataArray[:dataPTS]
        ymax = np.amax(QSubSet)
        ymin = np.amin(QSubSet)
        if rs_xlimit != []:
            xmin, xmax = rs_xlimit
        if rs_ylimit != []:
            ymin, ymax = rs_ylimit

        xtext = xmin + 0.3333 * (xmax - xmin)
        ytext1 = ymin + 0.5 * (ymax - ymin)
        ytext2 = ymin + 0.35 * (ymax - ymin)
        ytext3 = ymin + 0.2 * (ymax - ymin)
        ytext4 = ymin + 0.05 * (ymax - ymin)
        plt.text(xtext, ytext1, r'Oscillatory Frequency:  %.5f Hz ' % (frequency), color='blue',fontsize=14)
        plt.text(xtext, ytext2, r'Damping ratio (Logarithmic decrement): %.5f' % (damping_ratio), color='blue',fontsize=14)
        plt.text(xtext, ytext3, r'Halving Time:  %.5f Sec' % (halving_time), color='blue', fontsize=14)
        if (frequency < 0.05 and damping_ratio >=0.4) or (0.05 <= frequency <= 0.6 and halving_time<5) or (frequency>0.6 and damping_ratio>0.05):
            plt.text(xtext, ytext4, r'The Oscillation is adequately damped', color='blue', fontsize=14)
        else:
            plt.text(xtext, ytext4, r'The Oscillation is NOT adequately damped', color='red', fontsize=14)

    def Iq_plot(self,tDataArray,V_POC, Q_POC,Q_base,tFault,dProp_IqCap,dProp_IqInd):
        import matplotlib.pyplot as plt
        INV_VolData = V_POC
        INV_QData = Q_POC
        INV_QData = np.divide(INV_QData,Q_base) # INV_QData here is in MVAr, remove it if INV_QData in pu.
        
        #INV_VolData = V_INV # this value can be change to change the access point, between V_INV and V_POC, in  pu.
        #INV_QData = Q_INV # this value can be change to change the access point, between Q_INV and Q_POC in Mvar.

        dNER_IqCap = 4 #Automatic Standard value: 4% change of Iq for 1% change in V: this one to plot what the Iq should be according to V during the faults
        dProp_IqCap = dProp_IqCap #The Propose value in GPS: 1.5% change for LVRT as per proposed in GPS
        dNER_IqInd = 6 #Automatic Standard value: 6% for HVRT ?
        dProp_IqInd = dProp_IqInd #The Propose value in GPS: 0% change for HVRT as per proposed in GPS

        tFault = tFault   #time when you apply fault, important.
        IqCHL = np.divide(INV_QData,INV_VolData)
        plt.plot(tDataArray,IqCHL,'--',label="Actual_Iq")
        plt.autoscale(enable=True, axis='both', tight=False)
        simt = 0
        tDataArrayIndex = 0
        while simt < tFault:
            simt = tDataArray[tDataArrayIndex]
            tDataArrayIndex += 1
        #Reading INV V and Q 5 data points before tFault
        INV_V_PreFault = INV_VolData[tDataArrayIndex-5]
        INV_Q_PreFault = INV_QData[tDataArrayIndex-5]
        INV_Iq_PreFault = INV_Q_PreFault/INV_V_PreFault

        # Find Iq requirement during the fault as per NER
        Iq_NER = []
        Iq_Proposed = []
        tDataArrayIndex = 0
        for t in tDataArray:
            INV_V = INV_VolData[tDataArrayIndex]
            if INV_V < 0.85:
                Iq_ExpNER = INV_Iq_PreFault+(dNER_IqCap*(INV_V_PreFault-INV_V))
                Iq_ExpProp = INV_Iq_PreFault+(dProp_IqCap*(INV_V_PreFault-INV_V))
                if abs(Iq_ExpNER)>1:
                    Iq_ExpNER = Iq_ExpNER/abs(Iq_ExpNER)
                if abs(Iq_ExpProp)>1:
                    Iq_ExpProp = Iq_ExpProp/abs(Iq_ExpProp)
                Iq_NER.append(Iq_ExpNER)
                Iq_Proposed.append(Iq_ExpProp)
            elif INV_V > 1.15:
                Iq_ExpNER = INV_Iq_PreFault+(dNER_IqInd*(INV_V_PreFault-INV_V))
                Iq_ExpProp = INV_Iq_PreFault+(dProp_IqInd*(INV_V_PreFault-INV_V))
                if abs(Iq_ExpNER)>1:
                    Iq_ExpNER = Iq_ExpNER/abs(Iq_ExpNER)
                if abs(Iq_ExpProp)>1:
                    Iq_ExpProp = Iq_ExpProp/abs(Iq_ExpProp)
                Iq_NER.append(Iq_ExpNER)
                Iq_Proposed.append(Iq_ExpProp) # Have to update this one
            else:
                Iq_NER.append(INV_Iq_PreFault)
                Iq_Proposed.append(INV_Iq_PreFault)
            tDataArrayIndex += 1
        plt.plot(tDataArray,Iq_NER,':',label="NER_Auto_Iq")
        plt.plot(tDataArray,Iq_Proposed,'k:',label="Proposed_Iq")
        plt.autoscale(enable=True, axis='both', tight=False)

    def Vdroop_adjusted_plot(self,tDataArray,POC_V_ref,Q_POC, QDROOP_over_percent,QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower,deadband_over):
        import matplotlib.pyplot as plt
        QBASE = QBASE_pu * wfbase_MW
        QDROOP_over_pu = QDROOP_over_percent *0.01
        QDROOP_lower_pu = QDROOP_lower_percent *0.01
        V_Droop_error_over = np.divide((np.multiply(Q_POC, QDROOP_over_pu)), QBASE)
        V_Droop_error_lower = np.divide((np.multiply(Q_POC, QDROOP_lower_pu)), QBASE)
        V_Droop_Adjusted = np.subtract(POC_V_ref,V_Droop_error_over)
        V_Droop_Adjusted_over = np.subtract(POC_V_ref,V_Droop_error_over)
        V_Droop_Adjusted_lower = np.subtract(POC_V_ref, V_Droop_error_lower)

        for i in range(len(Q_POC)):
            if Q_POC[i] > 0.01:
                V_Droop_Adjusted[i] = V_Droop_Adjusted_over[i] - deadband_over
            elif Q_POC[i] < -0.01:
                V_Droop_Adjusted[i] = V_Droop_Adjusted_lower[i] + deadband_lower
            else:
                V_Droop_Adjusted[i] = V_Droop_Adjusted[i-1]

        V_Droop_Adjusted_Hi = np.add(V_Droop_Adjusted, 0.005)
        V_Droop_Adjusted_Li = np.subtract(V_Droop_Adjusted, 0.005)
        plt.plot(tDataArray,V_Droop_Adjusted,'--',label="V_Droop_Adjusted",linewidth = 3.5)
        plt.plot(tDataArray, V_Droop_Adjusted_Li, ':', label="Lower Boundary", color='gray',linewidth = 2.5)
        plt.plot(tDataArray, V_Droop_Adjusted_Hi, ':', label="Higher Boundary", color='gray',linewidth = 2.5)
        plt.autoscale(enable=True, axis='both', tight=False)
    def Qdroop_plot(self,tDataArray,POC_V_ref,V_POC, QDROOP_over_percent, QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower, deadband_over, Qrefupperlimit, Qreflowerlimit):
        import matplotlib.pyplot as plt
        QBASE = QBASE_pu * wfbase_MW
        QDROOP_over_pu = QDROOP_over_percent *0.01
        QDROOP_lower_pu = QDROOP_lower_percent *0.01
        QPCC = np.subtract(POC_V_ref,V_POC)
        for i in range(len(QPCC)):
            if QPCC[i] <=-deadband_over:
                QPCC[i] = max((QPCC[i] + deadband_over)*QBASE/(QDROOP_over_pu),-Qreflowerlimit*wfbase_MW)
            #lower
            elif QPCC[i] >= deadband_lower:
                QPCC[i] = min((QPCC[i] - deadband_lower)*QBASE/(QDROOP_lower_pu),Qrefupperlimit*wfbase_MW)
            else:
                QPCC[i] = 0
        plt.plot(tDataArray,QPCC,'--',label="Q_Droop_ref",linewidth = 3.5)
        plt.autoscale(enable=True, axis='both', tight=False)

    def Vdroop_Hybrid_adjusted_plot(self,tDataArray,POC_V_ref,POC_V_err,Q_POC, QDROOP, QBASE):
        import matplotlib.pyplot as plt
        QDROOP = QDROOP * 0.01
        V_Droop_error = np.divide((np.multiply(Q_POC,QDROOP)),QBASE)
        V_Adjusted = np.subtract(POC_V_ref,POC_V_err)
        V_Droop_Hybrid_Adjusted = np.subtract(V_Adjusted, V_Droop_error)
        plt.plot(tDataArray,V_Droop_Hybrid_Adjusted,'--',label="V_Droop_Hybrid_Adjusted",linewidth = 3.5)
        plt.autoscale(enable=True, axis='both', tight=False)

    def P_Recovery_Calc(self,tDataArray,INV_PData,tFault,clearnear,clearremote,filename,P_recovery_sheet):
        #from 100 milliseconds after clearance of the fault, active power of at least 95% of the level existing just prior to the fault.
        import matplotlib.pyplot as plt
        dataPTS = len(tDataArray)
        tFault = tFault #time to apply the fault
        Pmax = max(INV_PData)
        Pmin = min(INV_PData)
        #Find pre disturbance P
        simt = 0
        tDataArrayIndex = 0
        while simt < tFault:
            simt = tDataArray[tDataArrayIndex]
            tDataArrayIndex += 1
        #Reading INV V and Q 1 data points before tFault
        INV_P_PreFault = INV_PData[tDataArrayIndex-1]
        plt.plot([tDataArray[0],tDataArray[dataPTS-1]],[INV_P_PreFault*0.95,INV_P_PreFault*0.95],'r--',label="0.95xP_PreFAULT") #plot a horizental line at 0.95P pre fault
        plt.autoscale(enable=True, axis='both', tight=False)
        ymin, ymax = plt.ylim()
        tFClearN = clearnear + tFault
        tFClearR = clearremote + tFault
        tFClear = max(tFClearN,tFClearR)
        plt.plot([(tFClear+0.1),(tFClear+0.1)],[ymin,ymax],'r:') #plot vertical line at 100ms after clerance fault
        plt.autoscale(enable=True, axis='both', tight=False)
        plt.ylim(ymin,ymax) # This was inserted to reset ylim after previous plotting commands
        while simt < tFClear+0.1: #after 100ms following clearing the fault
            simt = tDataArray[tDataArrayIndex]
            tDataArrayIndex += 1
        #Reading P output after 100ms following clearing the fault
        INV_P_PostFault = INV_PData[tDataArrayIndex]    #P value after 100ms fault

        #if INV_P_PostFault >= INV_P_PreFault*0.95:
        #    plt.text((simt+0.6), (ymax+ymin)/2, r'P_recovery > 0.95xP_pre-flt',color='blue',fontsize=9)
        #else:
        #    plt.text((simt+0.6), (ymax+ymin)/2, r'P_recovery < 0.95xP_pre-flt',color='red',fontsize=9)

        simt = 0
        tDataArrayIndex = 0
        while simt < tFClear:
            simt = tDataArray[tDataArrayIndex]
            tDataArrayIndex += 1

        x = 0
        rec = 0
        while x <= INV_P_PreFault*0.95:
            x = INV_PData[tDataArrayIndex]
            tDataArrayIndex += 1
            if tDataArrayIndex >len(INV_PData)-1:
                rec = 1
                break
        if rec == 0:
            T_95 = tDataArray[tDataArrayIndex]
            T_recovery = (T_95 - tFClear)*1000
            if INV_P_PostFault >= INV_P_PreFault*0.95:
                plt.text((simt+0.6), (ymax+ymin)/2, r'P_recovery > 0.95xP_pre-flt',color='blue',fontsize=14)
                plt.text((simt+0.6), (ymax+ymin)/4, r'Recovery time = %5.2f (ms)'%T_recovery,color='blue',fontsize=14)
            else:
                plt.text((simt+0.6), (ymax+ymin)/2, r'P_recovery < 0.95xP_pre-flt',color='red',fontsize=14)
                plt.text((simt+0.6), (ymax+ymin)/4, r'Recovery time = %5.2f (ms)'%T_recovery,color='red',fontsize=14)
        else:
            T_recovery = 'Not able to recovery to 95% Pre-Fault Value'
            plt.text((simt + 0.6), (ymax + ymin) / 2, r'NOT ABLE TO RECOVERY to 95%', color='red', fontsize=14)
        from openpyxl import Workbook, load_workbook
        wb = load_workbook(P_recovery_sheet)
        ws = wb.active
        row = ws.max_row + 1
        ws.cell(row=row, column=1).value = os.path.basename(os.path.dirname(filename))
        ws.cell(row=row, column=2).value = T_recovery
        wb.save(P_recovery_sheet)

    def read_data(self, filename):
        # Read PSSE data

        if (filename[-4:].lower() in '.out') and ('pscad' not in filename.lower()):
            print('Plotting file: ' + str(filename))
            self.filenames.append(filename)
            # PSSE output fileslen(time)
            import os, psse34
            import dyntools

            outfile_data = dyntools.CHNF(filename)
            short_title, chanid, chandata = outfile_data.get_data()
            
            time = np.array(chandata['time'])
            data = np.zeros((len(chandata) - 1, len(time)))
            
            chan_ids = []
            for key in chandata.keys()[:-1]:
                print(key, chanid[key])
                data[key - 1] = chandata[key]
                chan_ids.append(chanid[key])
            self.timearrays.append(time)
            self.dataarrays.append(data)
            self.scales.append(np.ones(len(data)+1))
            self.offsets.append(np.zeros(len(data)+1))
            self.channel_names = chan_ids
        #Read PSCAD file
        if filename[-4:].lower() in ('*.inf'):
            print('Plotting file: ' + str(filename))
            # PSCAD output files
            import os, glob, re
            cwd = os.getcwd() 
            outfiles = []
            for pscadfile in glob.glob(filename[0:-4] + "_*.out"):
                outfiles.append(pscadfile)
            #
            chan_ids = []
            f = open(filename, 'r')
            lines = f.readlines()
            f.close()
            for idx, rawline in enumerate(lines):
                line = re.split(r'\s{1,}', rawline)
                chan_ids.append(line[2][6:-1])
            #
            # Setup time array using last outfile
            time = []
            f = open(outfiles[0], 'r')
            lines = f.readlines()
            f.close()
            for l_idx, rawline in enumerate(lines[1:]):
                line = re.split(r'\s{1,}', rawline)
                time.append(float(line[1]))
            time = np.array(time)
            #
            # Setup empty data array using known number of channels and time steps
            data = np.zeros((len(chan_ids), len(time)))
            # Populate data array using all outfiles
            for f_idx, outfile in enumerate(outfiles):
                f = open(outfile, 'r')
                lines = f.readlines()
                f.close()
                for n, rawline in enumerate(lines[1:]):
                    line = re.split(r'\s{1,}', rawline)
                    # print(line[1:-1])
                    for m, value in enumerate(line[2:-1]):
                        data[(f_idx*10) + m][n] = value
                        # data[(f_idx * 10) + 1 + m][n] = value
                        # print('Channel', (f_idx * 10) + 1 + m)
                    # data[0][n] = line[1]
            for n, chan_id in enumerate(chan_ids):
                print(n+1, chan_id)
                
            # Add data to arrays in class
            self.timearrays.append(time)
            self.dataarrays.append(data)
            self.scales.append(np.ones(len(data)+1))
            self.offsets.append(np.zeros(len(data)+1))
            self.channel_names = chan_ids

            # Read PSCAD INFX file
        if filename[-5:].lower() in ('*.infx'):
            print('Plotting file: ' + str(filename))
            # PSCAD output files
            import os, glob, re
            cwd = os.getcwd()
            outfiles = []
            for pscadfile in glob.glob(cwd + "/" + filename[0:-5] + "_*.out"):
                outfiles.append(pscadfile)
            #
            chan_ids = []
            f = open(filename, 'r')
            lines = f.readlines()
            f.close()
            lines = lines[5:(len(lines) - 3)]
            for idx, rawline in enumerate(lines):
                name_start = rawline.find("Analog name=") + 13
                name_stop = rawline.find("index") - 2
                rawline_short = rawline[name_start:name_stop]
                name_start = rawline_short.find(":") + 1
                chan_id = rawline_short[name_start:]

                name_start = rawline.find("dim=") + 5
                name_stop = rawline.find("unit") - 2
                dim = rawline[name_start:name_stop]
                dim = int(dim)
                if dim == 1:
                    chan_ids.append(chan_id)
                else:
                    for i in range(dim):
                        chan_ids.append(chan_id + '_%d' %(i+1))
            print('chan_ids:')
            print(len(chan_ids))
            #
            # Setup time array using last outfile
            time = []
            f = open(outfiles[0], 'r')
            lines = f.readlines()
            f.close()
            for l_idx, rawline in enumerate(lines[1:]):
                line = re.split(r'\s{1,}', rawline)
                time.append(float(line[1]))
            time = np.array(time)
            #
            # Setup empty data array using known number of channels and time steps
            data = np.zeros((len(chan_ids), len(time)))
            #
            # Populate data array using all outfiles
            for f_idx, outfile in enumerate(outfiles):
                f = open(outfile, 'r')
                lines = f.readlines()
                f.close()
                for n, rawline in enumerate(lines[1:]):
                    line = re.split(r'\s{1,}', rawline)
                    for m, value in enumerate(line[2:-1]):
                        data[(f_idx * 10) + m][n] = value
            for n, chan_id in enumerate(chan_ids):
                print(n + 1, chan_id)

            # Add data to arrays in class
            self.timearrays.append(time)
            self.dataarrays.append(data)
            self.scales.append(np.ones(len(data) + 1))
            self.offsets.append(np.zeros(len(data) + 1))
            self.channel_names = chan_ids

        #Read Excel PSSE standard file
        if filename[-5:].lower() in ('*.xlsx'):
            print('Plotting file: ' + str(filename))
            import os
            from os.path import join, dirname, abspath, isfile
            from collections import Counter
            import xlrd
            from xlrd.sheet import ctype_text
            cwd = os.getcwd()
            import psse34
            import dyntools
            def read_excel_tools(filename,sheet_index):
                xl_workbook = xlrd.open_workbook(filename)
                xl_sheet = xl_workbook.sheet_by_index(sheet_index)
                return xl_sheet
            def get_data(xl_sheet):
                chanid = {}
                chandata = {}
                row = xl_sheet.row(2)
                short_title = str(xl_sheet.row(1)[0].value) + '\n' + str(xl_sheet.row(2)[0].value)

                row_chan = xl_sheet.row(3)
                row_title = xl_sheet.row(4)
                for chan_idx in range(len(row_chan)):
                    chan_idx_data = []
                    for row_idx in range(5, xl_sheet.nrows):
                        cell_obj = xl_sheet.cell(row_idx, chan_idx)
                        chan_idx_data.append(cell_obj.value)
                    if chan_idx ==0:
                        chanid['time'] = str(row_title[chan_idx].value)
                        chandata['time'] = chan_idx_data
                    else:
                        chanid[int(row_chan[chan_idx].value)] = str(row_title[chan_idx].value)
                        chandata[int(row_chan[chan_idx].value)] = chan_idx_data
                return short_title, chanid, chandata

            xl_sheet = read_excel_tools(filename,sheet_index = 0)
            short_title, chanid, chandata = get_data(xl_sheet)
            time = np.array(chandata['time'])
            data = np.zeros((len(chandata) - 1, len(time)))
            chan_ids = []
            for key in chandata.keys()[:-1]:
                print(key, chanid[key])
                data[key - 1] = chandata[key]
                chan_ids.append(chanid[key])
            print('\n')
            self.timearrays.append(time)
            self.dataarrays.append(data)
            self.scales.append(np.ones(len(data)+1))
            self.offsets.append(np.zeros(len(data)+1))
            self.channel_names = chan_ids

    def read_data2(self, filename, sheet_index_or_name):
        #Read Excel file
        if filename[-5:].lower() in ('*.xlsx'):
            print('Plotting file: ' + str(filename))
            import os
            from os.path import join, dirname, abspath, isfile
            from collections import Counter
            import xlrd
            from xlrd.sheet import ctype_text
            import psse34
            import dyntools
            def read_excel_tools(filename, sheet_index_or_name):
                xl_workbook = xlrd.open_workbook(filename)
                if type(sheet_index_or_name) == int:
                    xl_sheet = xl_workbook.sheet_by_index(sheet_index_or_name)
                elif type(sheet_index_or_name) == str:
                    xl_sheet = xl_workbook.sheet_by_name(sheet_index_or_name)
                return xl_sheet
            def get_data(xl_sheet):
                import re
                chanid = {}
                chandata = {}
                chan_number_row = 0
                title_row = 0
                short_title = os.path.splitext(filename)[0]
                row_chan = xl_sheet.row(chan_number_row)
                row_title = xl_sheet.row(title_row)
                for chan_idx in range(len(row_chan)):
                    chan_idx_data = []
                    for row_idx in range(1, xl_sheet.nrows):
                        cell_obj = (xl_sheet.cell(row_idx, chan_idx)).value
                        if (isinstance(cell_obj, unicode)):
                            if len(re.findall(("\d+\.\d+"), cell_obj)) >0:
                                cell_obj = float(re.findall("\d+\.\d+", cell_obj)[0])
                            elif len(re.findall(("\d+"), cell_obj)) > 0:
                                cell_obj = float(re.findall("\d+", cell_obj)[0])
                            else:
                                cell_obj = 0
                        #print(cell_obj)
                        # if type(cell_obj) != float:
                        #     cell_obj = float(re.findall("\d+\.\d+", cell_obj)[0])
                        chan_idx_data.append(cell_obj)
                    if chan_idx == 0:
                        chanid['time'] = str(row_title[chan_idx].value)
                        chandata['time'] = chan_idx_data
                    else:
                        chanid[int(chan_idx)] = str(row_title[chan_idx].value)
                        chandata[int(chan_idx)] = chan_idx_data
                return short_title, chanid, chandata
            xl_sheet = read_excel_tools(filename,sheet_index_or_name = sheet_index_or_name)
            short_title, chanid, chandata = get_data(xl_sheet)
            time = np.array(chandata['time'])
            data = np.zeros((len(chandata) - 1, len(time)))
            chan_ids = []
            for key in chandata.keys()[:-1]:
                print(key, chanid[key])
                data[key - 1] = chandata[key]
                chan_ids.append(chanid[key])
            print('\n')
            self.timearrays.append(time)
            self.dataarrays.append(data)
            self.scales.append(np.ones(len(data)+1))
            self.offsets.append(np.zeros(len(data)+1))
            self.channel_names = chan_ids

    def subplot_spec(self, subplot, plot_arrays, title = '', ylabel = '',xlimit='', ylimit ='', scale = 1.0, offset = 0.0, rstime ='',plot_iq ='',
                     P_Recovery='', timeoffset='',plot_Vdroop ='',plot_Qdroop ='',plot_Vdroops_Hybrid ='',plot_PF ='',checkspike ='', damping_cal=''):
        """
        Specify which channels are to be plotted

        Args:
            subplot: Number of subplot for which we are specifying files 
            and channels
            
            plot_arrays: The plot_arrays input should be specified as a two element tuple, 
            with the first element being the input file and the second being the channel 
            of that file to be plotted. The second element can be specified either 
            as the channel number (as an int), or the name as shown in 
            self.channel_names (as a str). If a string is provided, only the first
            few letters need to be specified sufficient to be unique from the other
            channel_names. If the string is not unique then the first match will be 
            plotted.

            title: Subplot title (optional)
            
            ylabel: Y axis title (optional)
            
            scale: Constant multiplier for this trace (optional)
            
            offset: Y axis offset for this trace (optional)

        Returns:
            Nothing

        Raises:
            Nothing

        """
        if plot_arrays[0] < len(self.dataarrays):
            if title != '':
                self.titles[subplot] = title
            if ylabel != '':
                self.ylabels[subplot] = ylabel
            if ylimit != '':
                self.ylimits[subplot] = ylimit
            if xlimit != '':
                self.xlimit = xlimit
            if rstime != '':
                self.rstimes[subplot] = rstime
            if damping_cal != '':
                self.damping_cals[subplot] = damping_cal
            if plot_iq != '':
                self.plot_iqs[subplot] = plot_iq
            if plot_Vdroop != '':
                self.plot_Vdroops[subplot] = plot_Vdroop
            if plot_Qdroop != '':
                self.plot_Qdroops[subplot] = plot_Qdroop
            if plot_Vdroops_Hybrid != '':
                self.plot_Vdroops_Hybrids[subplot] = plot_Vdroops_Hybrid

            if P_Recovery != '':
                self.P_Recoverys[subplot] = P_Recovery

            if timeoffset != '':
                self.timeoffset = [timeoffset]

            if plot_PF != '':
                self.plot_PFs[subplot] = plot_PF
            if checkspike != '':
                self.checkspikes[subplot] = checkspike


            #
            # if type(plot_arrays[1]) is str:  #self.channel_names
            #     chars = len(plot_arrays[1])
            #     for idx, name in enumerate(self.channel_names[plot_arrays[0]]):
            #         if name[:chars] == plot_arrays[1]:
            #             plot_channel = idx + 1  # Channel numbers start at 1
            #             break

            if type(plot_arrays[1]) is str:  #self.channel_names
                for n, chan_id in enumerate(self.channel_names):
                    if plot_arrays[1].lower() == self.channel_names[n].lower():
                        plot_channel = n+1 # Channel numbers start at 1
                        break
            else:
                plot_channel = plot_arrays[1]
            #
            self.plotspec[subplot].append((plot_arrays[0], plot_channel))
            self.scales[plot_arrays[0]][plot_channel] = scale
            self.offsets[plot_arrays[0]][plot_channel] = offset
        else:
            print('Specified file number {:d} is not in memory'.format(plot_arrays[0]))

            
    def plot(self, figname = '', show = 0):
        '''
        Create plot of channels and files as specified in self.plotspec.\n
        Manually specify the following plot variables as required:\n
            self.legends[subplot]\n
            self.titles[subplot]\n
            self.ylabels[subplot]\n
            self.ylimit[subplot]\n
            self.xlimit
        '''
        import matplotlib.pyplot as plt

        if show:
            plt.ion() #Turn the interactive mode on.
            plt.pause(0.0001)

        from matplotlib.ticker import ScalarFormatter, FormatStrFormatter

        for idx, i in enumerate(self.plotspec):
            if i != []:
                subplots = idx + 1
        
        if subplots == 1:
            subplot_index = [111]
            nRows = 1
            nCols = 1
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 18 * nCols), (2 + 10 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98, left = 0.1, hspace = 0.25)
        if subplots == 2:
            subplot_index = [211, 212]
            nRows = 2
            nCols = 1
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 18 * nCols), (2 + 5 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98, left = 0.1, hspace = 0.28)
        if subplots == 3:
            subplot_index = [311, 312, 313]
            nRows = 3
            nCols = 1
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 24 * nCols), (2 + 5 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98,  left = 0.1, hspace = 0.32, wspace = 0.2)
        if subplots == 4:
            nRows = 2
            nCols = 2
            subplot_index = [221, 222, 223, 224]
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 12 * nCols), (2 + 6.5 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98, left = 0.1, hspace = 0.32, wspace = 0.2)
        if subplots == 5:
            subplot_index = [311, 323, 324, 325, 326]
            nRows = 3
            nCols = 2
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 12 * nCols), (2 + 5 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98, left = 0.1, hspace = 0.35, wspace = 0.2)
        if subplots == 6:
            nRows = 3
            nCols = 2
            subplot_index = [321, 322, 323, 324, 325, 326]
            #plt.figure(figsize = (29.7/ 2.54, 21.0/ 2.54))
            plt.figure(figsize=[(2 + 12 * nCols), (2 + 5 * nRows)])
            #plt.subplots_adjust(bottom = 0.05, top = 0.91, right = 0.98, left = 0.1, hspace = 0.37, wspace = 0.2)
        import os
        plot_title = os.path.basename(figname)
        # plot_title = os.path.basename(os.path.dirname(figname))

        plt.suptitle(plot_title,fontsize = 22,fontweight ='bold')
        
        #
        plt.rc('xtick', labelsize = 18)
        plt.rc('ytick', labelsize = 18)
        
        #print 'Number of subplots: ' + str(subplots)
        #print 'Detail plots:' + str(self.plotspec)
        
        for idx in range(subplots):
            plt.subplot(subplot_index[idx])
            ax = plt.gca()
            # ax.yaxis.set_major_formatter(FormatStrFormatter('%1.{:d}f'.format(2)))            
            ax.ticklabel_format(useOffset=False)

            legend_chan_names = []
            for trace in self.plotspec[idx]:
                datafile = trace[0]
                channel = trace[1]
                if channel != 0:
                    plt.plot(self.timearrays[datafile] + self.timeoffset[datafile],
                                                        (self.dataarrays[datafile][channel-1]
                                                        * self.scales[datafile][channel])
                                                        + self.offsets[datafile][channel], linewidth = 3.5)

                chan_name =  self.channel_names[channel-1]
                legend_chan_names.append(chan_name)

            #Plot raise time and settling time
            #trace = (self.plotspec)[idx][len(self.plotspec[idx])-1] #only plot settling time rise time at last
                

            #Plot Iq
            trace = (self.plotspec)[idx][0]  #only plot iq at first              
            datafile = trace[0]
            channel = trace[1]
            if self.plot_iqs[idx] != []:
                V_POC_Channel,Q_POC_Channel, Q_base, tFault,dProp_IqCap,dProp_IqInd = self.plot_iqs[idx]

                if isinstance(V_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if V_POC_Channel.lower() == self.channel_names[n].lower():
                            V_POC_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    V_POC_Channel_id = V_POC_Channel

                if isinstance(Q_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if Q_POC_Channel.lower() == self.channel_names[n].lower():
                            Q_POC_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    Q_POC_Channel_id = Q_POC_Channel
                self.Iq_plot(self.timearrays[datafile] + self.timeoffset[datafile],self.dataarrays[datafile][V_POC_Channel_id-1],
                             self.dataarrays[datafile][Q_POC_Channel_id-1],Q_base, tFault,dProp_IqCap,dProp_IqInd)

            if self.plot_Vdroops[idx] != []:
                V_POC_Channel, Q_POC_Channel, QDROOP_over_percent, QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower, deadband_over= self.plot_Vdroops[idx]

                if isinstance(V_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if V_POC_Channel.lower() == self.channel_names[n].lower():
                            POC_V_ref_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    POC_V_ref_Channel_id = V_POC_Channel

                if isinstance(Q_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if Q_POC_Channel.lower() == self.channel_names[n].lower():
                            Q_POC_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    Q_POC_Channel_id = Q_POC_Channel

                self.Vdroop_adjusted_plot(self.timearrays[datafile] + self.timeoffset[datafile],
                                          self.dataarrays[datafile][POC_V_ref_Channel_id - 1],
                                          self.dataarrays[datafile][Q_POC_Channel_id - 1],QDROOP_over_percent, QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower, deadband_over)

            if self.plot_Qdroops[idx] != []:
                V_ref_Channel,V_POC_Channel, QDROOP_over_percent, QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower, deadband_over, Qrefupperlimit, Qreflowerlimit = self.plot_Qdroops[idx]

                if isinstance(V_ref_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if V_ref_Channel.lower() == self.channel_names[n].lower():
                            POC_V_ref_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    POC_V_ref_Channel_id = V_ref_Channel

                if isinstance(V_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if V_POC_Channel.lower() == self.channel_names[n].lower():
                            V_POC_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    V_POC_Channel_id = V_POC_Channel

                self.Qdroop_plot(self.timearrays[datafile] + self.timeoffset[datafile],
                                          self.dataarrays[datafile][POC_V_ref_Channel_id - 1],
                                          self.dataarrays[datafile][V_POC_Channel_id - 1],QDROOP_over_percent, QDROOP_lower_percent, QBASE_pu, wfbase_MW, deadband_lower, deadband_over, Qrefupperlimit, Qreflowerlimit)

            if self.plot_Vdroops_Hybrids[idx] != []:
                POC_V_ref_Channel_id,V_err_Channel_id,Q_POC_Channel_id, QDROOP, QBASE = self.plot_Vdroops_Hybrids[idx]
                self.Vdroop_Hybrid_adjusted_plot(self.timearrays[datafile] + self.timeoffset[datafile],
                                                 self.dataarrays[datafile][POC_V_ref_Channel_id - 1],
                                                 self.dataarrays[datafile][V_err_Channel_id - 1],
                                                 self.dataarrays[datafile][Q_POC_Channel_id - 1],QDROOP,QBASE)


            #Plot P Recovery
            trace = (self.plotspec)[idx][0]  #only plot iq at first              
            datafile = trace[0]
            channel = trace[1]            
            if self.P_Recoverys[idx] != []:
                P_POC_Channel,tFault,clearnear,clearremote, filename,P_recovery_sheet = self.P_Recoverys[idx]

                if isinstance(P_POC_Channel, str):
                    for n, chan_id in enumerate(self.channel_names):
                        if P_POC_Channel.lower() == self.channel_names[n].lower():
                            P_POC_Channel_id = n + 1  # Channel numbers start at 1
                            break
                else:
                    P_POC_Channel_id = P_POC_Channel

                self.P_Recovery_Calc(self.timearrays[datafile] + self.timeoffset[datafile],self.dataarrays[datafile][P_POC_Channel_id-1],tFault,clearnear,clearremote, filename,P_recovery_sheet)
            # Plot PF
            trace = (self.plotspec)[idx][0]  # only plot iq at first
            datafile = trace[0]
            channel = trace[1]
            if self.plot_PFs[idx] != []:
                P_POC_Channel_id, Q_POC_Channel_id, PF_set_Channel_id = self.plot_PFs[idx]
                self.PF_plot(self.timearrays[datafile] + self.timeoffset[datafile],
                             self.dataarrays[datafile][P_POC_Channel_id - 1],
                             self.dataarrays[datafile][Q_POC_Channel_id - 1],
                             self.dataarrays[datafile][PF_set_Channel_id - 1])

            trace = (self.plotspec)[idx][0]  # only plot settling time rise time at first
            datafile = trace[0]
            channel = trace[1]

            rs_ylimit = self.ylimits[idx]
            rs_xlimit = self.xlimit

            # if self.rstimes[idx] != []:
            #     tFault, Tfin, filename,workbook = self.rstimes[idx]
            #     self.rise_settle_TimeCalc(self.timearrays[datafile] + self.timeoffset[datafile],
            #                               (self.dataarrays[datafile][channel - 1] * self.scales[datafile][channel]) +
            #                               self.offsets[datafile][channel],rs_xlimit,rs_ylimit, tFault, Tfin, filename,workbook)
            if self.rstimes[idx] != []:
                tFault, Tfin = self.rstimes[idx]
                self.rise_settle_TimeCalc(self.timearrays[datafile] + self.timeoffset[datafile],
                                          (self.dataarrays[datafile][channel - 1] * self.scales[datafile][channel]) +
                                          self.offsets[datafile][channel],rs_xlimit,rs_ylimit, tFault, Tfin)

            trace = (self.plotspec)[idx][0]  # only plot settling time rise time at first
            datafile = trace[0]
            channel = trace[1]
            rs_ylimit = self.ylimits[idx]
            rs_xlimit = self.xlimit
            if self.damping_cals[idx] != []:
                Parameter, value = self.damping_cals[idx]
                self.damping_calculation(self.timearrays[datafile] + self.timeoffset[datafile],
                                (self.dataarrays[datafile][channel - 1] * self.scales[datafile][channel]) + self.offsets[datafile][channel],rs_xlimit, rs_ylimit, Parameter, value)

            trace = (self.plotspec)[idx][0]  # only plot settling time rise time at first
            datafile = trace[0]
            channel = trace[1]
            if self.checkspikes[idx] != []:
                thereshold, type = self.checkspikes[idx]
                self.plot_spike(self.timearrays[datafile] + self.timeoffset[datafile],
                                (self.dataarrays[datafile][channel - 1] * self.scales[datafile][channel]) + self.offsets[datafile][channel], thereshold, type)


            '''top = 0.9
            if self.titles[idx] == "":
                top = 0.95
            plt.subplots_adjust(right=0.9, hspace=0.3, wspace=0.3, top=top, left=0.1, bottom=0.12)'''

            #if self.legends[idx] != []:
            top = 0.9
            if self.titles[idx] == "":
                top = 0.95
            plt.subplots_adjust(right=0.9, hspace=0.3, wspace=0.3, top=top, left=0.1, bottom=0.12)
            #plt.legend(self.legends[idx], prop = {'size': 16}, loc = 'center right')

            if self.legends[idx] != []:
                legend = plt.legend(self.legends[idx],bbox_to_anchor=(0., -0.25, 1., .105), loc=2,
                                    ncol=2, mode="expand", prop={'size': 16}, borderaxespad=0.1,frameon =False)
                legend.get_title().set_fontsize('16')  # legend 'Title' fontsize
                plt.setp(plt.gca().get_legend().get_texts(), fontsize='16')  # legend 'list' fontsize
            else:
                if self.plot_Vdroops[idx] != []:
                    legend_chan_names.append('V_Droop_Ref_Adjusted')
                    legend_chan_names.append('Lower Boundary')
                    legend_chan_names.append('Higher Boundary')
                    legend = plt.legend(legend_chan_names,bbox_to_anchor=(0., -0.25, 1., .105), loc=2,
                                        ncol=2, mode="expand", prop={'size': 16}, borderaxespad=0.1,frameon =False)
                if self.plot_Qdroops[idx] != []:
                    legend_chan_names.append('Q_Droop_Ref')
                    legend = plt.legend(legend_chan_names,bbox_to_anchor=(0., -0.25, 1., .105), loc=2,
                                        ncol=2, mode="expand", prop={'size': 16}, borderaxespad=0.1,frameon =False)

                elif self.plot_Vdroops_Hybrids[idx] != []:
                    legend_chan_names.append('Vdroops_ref_Hybrids_Adjusted')
                    legend = plt.legend(legend_chan_names,bbox_to_anchor=(0., -0.25, 1., .105), loc=2,
                                        ncol=2, mode="expand", prop={'size': 16}, borderaxespad=0.1,frameon =False)

                else:
                    legend = plt.legend(legend_chan_names,bbox_to_anchor=(0., -0.25, 1., .105), loc=2,
                                        ncol=2, mode="expand", prop={'size': 16}, borderaxespad=0.1,frameon =False)

            plt.subplots_adjust(hspace=0.5)

            if self.titles[idx] != []:
                plt.title(self.titles[idx], fontsize = 20,fontweight ='bold')
            #
            if self.ylabels[idx] != []:
                plt.ylabel(self.ylabels[idx], fontsize = 18)
                plt.ylabel(self.ylabels[idx], labelpad=20)
            #
            if self.ylimits[idx] != []:
                plt.ylim(self.ylimits[idx])
            #
            if self.xlimit != []:
                plt.xlim(self.xlimit)
            else:
                xmin, xmax = ax.get_xlim()
                ticks = ax.get_xticks()
                plt.xlim([0, ticks[-2]])

                #'xtick', labelsize=SMALL_SIZE
            plt.xlabel('Time (sec)', fontsize = 16)
            plt.grid(1)

            # plt.get_current_fig_manager().window.state('zoomed')
            # plt.set_size_inches(11.69, 8.27)  # A4
            # tight_layout(pad=1.08, h_pad=None, w_pad=None, rect=None)[source]
            # plt.tight_layout(pad=1.08, h_pad=2, w_pad=2, rect=(0, 0, 1.0, 0.95))


        if figname != '':
            #plt.tight_layout() #pad=0.4, w_pad=0.5, h_pad=1.0)
            # plt.savefig(figname + '.png', format = 'png')
            plt.savefig(figname + '.pdf', format = 'pdf')
            plt.close()
            #plt.clf()
            #plt.savefig(figname + '.emf', format = 'emf')
            #plt.savefig(figname + '.svg', format = 'svg')
            #plt.savefig(figname + '.eps', format='eps')
        if show:
            plt.show(block=True)
